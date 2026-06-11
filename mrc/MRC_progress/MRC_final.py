#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import argparse
import glob
import json
import math
import os
from typing import Any, Dict, List, Optional, Tuple

import cv2
import matplotlib
import numpy as np
import openpyxl

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager
from matplotlib.ticker import FuncFormatter

RectInfo = Dict[str, Any]


class MRCFinalProcessor:
    def __init__(self, mapping_table_path: str) -> None:
        self.mapping_grid_10x10 = self._load_mapping_grid_10x10(mapping_table_path)
        self.min_area = 180.0
        self.max_area = 1800.0
        self.max_aspect_ratio = 1.9
        # 仅用于「四角波峰波谷定向」的宽松阈值（大角度/旋转后仍尽量凑够角点候选）
        self._corner_orient_min_area = 90.0
        self._corner_orient_max_area = 2800.0
        self._corner_orient_max_aspect_ratio = 2.8

    @staticmethod
    def _imread_unicode(path: str) -> Optional[np.ndarray]:
        try:
            data = np.fromfile(path, dtype=np.uint8)
            if data.size == 0:
                return None
            return cv2.imdecode(data, cv2.IMREAD_COLOR)
        except Exception:
            return None

    @staticmethod
    def _imwrite_unicode(path: str, image: np.ndarray) -> bool:
        ext = os.path.splitext(path)[1] or ".png"
        ok, buf = cv2.imencode(ext, image)
        if not ok:
            return False
        try:
            buf.tofile(path)
            return True
        except Exception:
            return False

    @staticmethod
    def _load_mapping_grid_10x10(mapping_path: str) -> List[List[str]]:
        if not os.path.exists(mapping_path):
            raise FileNotFoundError(f"未找到映射表文件: {mapping_path}")
        wb = openpyxl.load_workbook(mapping_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        grid: List[List[str]] = []
        for r in range(1, 11):
            row_vals: List[str] = []
            for c in range(1, 11):
                v = ws.cell(row=r, column=c).value
                if v is None:
                    raise ValueError(f"映射表存在空单元格: r{r} c{c}")
                row_vals.append(str(v).strip().replace("_", "-"))
            grid.append(row_vals)
        return grid

    # ------------ rotation detection + deskew (from ocr.py) ------------
    @staticmethod
    def _build_gray_preview(image_bgr: np.ndarray) -> np.ndarray:
        gray = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2GRAY)
        clahe = cv2.createCLAHE(clipLimit=1.2, tileGridSize=(8, 8))
        local = clahe.apply(gray)
        background = cv2.GaussianBlur(local, (0, 0), 5.0)
        detail = cv2.subtract(local, background)
        detail = cv2.normalize(detail, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
        return cv2.addWeighted(local, 0.82, detail, 0.18, 0.0)

    @staticmethod
    def _build_structure_mask(image_bgr: np.ndarray) -> np.ndarray:
        gray_local = MRCFinalProcessor._build_gray_preview(image_bgr)
        background = cv2.GaussianBlur(gray_local, (0, 0), 11.0)
        highpass = cv2.subtract(gray_local, background)
        highpass = cv2.normalize(highpass, None, 0, 255, cv2.NORM_MINMAX)

        tophat = cv2.morphologyEx(
            gray_local,
            cv2.MORPH_TOPHAT,
            cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (19, 19)),
        )

        grad_x = cv2.Sobel(gray_local, cv2.CV_32F, 1, 0, ksize=3)
        grad_y = cv2.Sobel(gray_local, cv2.CV_32F, 0, 1, ksize=3)
        grad_mag = cv2.magnitude(grad_x, grad_y)
        grad_mag = cv2.GaussianBlur(grad_mag, (3, 3), 0)
        grad_mag = cv2.normalize(grad_mag, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)

        _, m1 = cv2.threshold(highpass.astype(np.uint8), 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        _, m2 = cv2.threshold(tophat, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        _, m3 = cv2.threshold(grad_mag, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        mask = cv2.bitwise_or(cv2.bitwise_or(m1, m2), m3)
        mask = cv2.morphologyEx(
            mask,
            cv2.MORPH_OPEN,
            cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3)),
            iterations=1,
        )
        mask = cv2.morphologyEx(
            mask,
            cv2.MORPH_CLOSE,
            cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5)),
            iterations=1,
        )

        num_labels, labels, stats, _ = cv2.connectedComponentsWithStats(mask, connectivity=8)
        filtered = np.zeros_like(mask)
        h, w = mask.shape
        image_area = float(h * w)
        for idx in range(1, num_labels):
            area = int(stats[idx, cv2.CC_STAT_AREA])
            if area < 12 or area > int(image_area * 0.015):
                continue
            cw = int(stats[idx, cv2.CC_STAT_WIDTH])
            ch = int(stats[idx, cv2.CC_STAT_HEIGHT])
            if min(cw, ch) < 3:
                continue
            if max(cw, ch) / max(min(cw, ch), 1) > 6.0:
                continue
            filtered[labels == idx] = 255
        return filtered

    @staticmethod
    def _collect_central_points(mask: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
        ys, xs = np.where(mask > 0)
        if len(xs) < 100:
            raise RuntimeError("亮结构太少，无法建立参考矩形。")

        center_x = float(np.mean(xs))
        center_y = float(np.mean(ys))
        bbox_w = float(xs.max() - xs.min() + 1)
        bbox_h = float(ys.max() - ys.min() + 1)
        radius = 0.22 * max(bbox_w, bbox_h)

        num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(mask, connectivity=8)
        selected_mask = np.zeros_like(mask)
        points = []
        for idx in range(1, num_labels):
            if int(stats[idx, cv2.CC_STAT_AREA]) < 20:
                continue
            cx, cy = centroids[idx]
            if float(np.hypot(cx - center_x, cy - center_y)) > radius:
                continue
            ys_i, xs_i = np.where(labels == idx)
            if len(xs_i) == 0:
                continue
            selected_mask[ys_i, xs_i] = 255
            points.append(np.column_stack([xs_i.astype(np.float32), ys_i.astype(np.float32)]))

        if not points:
            raise RuntimeError("中心参考矩形提取失败。")
        return np.vstack(points), selected_mask

    @staticmethod
    def _normalize_angle_to_45(angle_deg: float) -> float:
        while angle_deg <= -45.0:
            angle_deg += 90.0
        while angle_deg > 45.0:
            angle_deg -= 90.0
        return angle_deg

    @staticmethod
    def _fit_reference_rectangle(points: np.ndarray) -> Tuple[np.ndarray, float]:
        hull = cv2.convexHull(points.reshape(-1, 1, 2))
        rect = cv2.minAreaRect(hull)
        box = cv2.boxPoints(rect).astype(np.float32)

        best_angle = None
        best_abs = 1e9
        for idx in range(4):
            p1 = box[idx]
            p2 = box[(idx + 1) % 4]
            angle = float(np.degrees(np.arctan2(float(p2[1] - p1[1]), float(p2[0] - p1[0]))))
            angle = MRCFinalProcessor._normalize_angle_to_45(angle)
            if abs(angle) < best_abs:
                best_abs = abs(angle)
                best_angle = angle

        if best_angle is None:
            raise RuntimeError("参考矩形角度计算失败。")
        return np.int32(np.round(box)), float(best_angle)

    @staticmethod
    def _measure_side_angle(image_bgr: np.ndarray) -> Tuple[float, np.ndarray, np.ndarray, np.ndarray]:
        mask = MRCFinalProcessor._build_structure_mask(image_bgr)
        points, selected_mask = MRCFinalProcessor._collect_central_points(mask)
        box, angle_deg = MRCFinalProcessor._fit_reference_rectangle(points)
        return angle_deg, mask, selected_mask, box

    @staticmethod
    def _estimate_dark_border_value(image: np.ndarray) -> Tuple[int, int, int]:
        h, w = image.shape[:2]
        bw = max(8, w // 30)
        bh = max(8, h // 30)
        strips = [
            image[:bh, :, :].reshape(-1, 3),
            image[-bh:, :, :].reshape(-1, 3),
            image[:, :bw, :].reshape(-1, 3),
            image[:, -bw:, :].reshape(-1, 3),
        ]
        pixels = np.vstack(strips)
        value = np.percentile(pixels, 15, axis=0)
        return tuple(int(np.clip(v, 0, 255)) for v in value)

    @staticmethod
    def _rotate_keep_size(image: np.ndarray, delta_deg: float, border_value: Tuple[int, int, int]) -> np.ndarray:
        h, w = image.shape[:2]
        matrix = cv2.getRotationMatrix2D((w / 2.0, h / 2.0), delta_deg, 1.0)
        return cv2.warpAffine(
            image,
            matrix,
            (w, h),
            flags=cv2.INTER_LINEAR,
            borderMode=cv2.BORDER_CONSTANT,
            borderValue=border_value,
        )

    @staticmethod
    def _rotate_keep_size_with_pad(
        image: np.ndarray, delta_deg: float, border_value: Tuple[int, int, int], pad_ratio: float = 0.18
    ) -> np.ndarray:
        """扩边后旋转再裁回原尺寸，减轻大角度时边缘裁切丢目标（用于 90° 定向等）。"""
        h, w = image.shape[:2]
        pad = int(round(max(h, w) * pad_ratio))
        padded = cv2.copyMakeBorder(image, pad, pad, pad, pad, cv2.BORDER_CONSTANT, value=border_value)
        ph, pw = padded.shape[:2]
        matrix = cv2.getRotationMatrix2D((pw / 2.0, ph / 2.0), delta_deg, 1.0)
        rotated = cv2.warpAffine(
            padded,
            matrix,
            (pw, ph),
            flags=cv2.INTER_LINEAR,
            borderMode=cv2.BORDER_CONSTANT,
            borderValue=border_value,
        )
        y1 = max(0, (ph - h) // 2)
        x1 = max(0, (pw - w) // 2)
        return rotated[y1 : y1 + h, x1 : x1 + w]

    def _refine_rotation_with_verification(
        self,
        image_bgr: np.ndarray,
        initial_angle_deg: float,
        border_value: Tuple[int, int, int],
        max_steps: int = 5,
        tolerance_deg: float = 0.35,
    ) -> Tuple[np.ndarray, float, float]:
        candidates = []
        for signed_angle in (-initial_angle_deg, initial_angle_deg):
            rotated = self._rotate_keep_size(image_bgr, signed_angle, border_value)
            try:
                residual_angle, _, _, _ = self._measure_side_angle(rotated)
            except Exception:
                residual_angle = 1e9
            candidates.append((abs(residual_angle), rotated, signed_angle, residual_angle))

        candidates.sort(key=lambda item: item[0])
        best_abs, best_image, applied_rotation_deg, residual = candidates[0]

        steps = 0
        while best_abs > tolerance_deg and steps < max_steps:
            correction = -residual
            updated = self._rotate_keep_size(best_image, correction, border_value)
            try:
                new_residual, _, _, _ = self._measure_side_angle(updated)
            except Exception:
                break
            if abs(new_residual) >= best_abs:
                break
            best_image = updated
            applied_rotation_deg += correction
            residual = new_residual
            best_abs = abs(new_residual)
            steps += 1

        if best_abs > tolerance_deg:
            fine_best = (best_abs, best_image, applied_rotation_deg, residual)
            for delta in np.linspace(-1.2, 1.2, 25):
                trial = self._rotate_keep_size(best_image, float(delta), border_value)
                try:
                    trial_residual, _, _, _ = self._measure_side_angle(trial)
                except Exception:
                    continue
                trial_abs = abs(float(trial_residual))
                if trial_abs < fine_best[0]:
                    fine_best = (trial_abs, trial, applied_rotation_deg + float(delta), float(trial_residual))
            best_abs, best_image, applied_rotation_deg, residual = fine_best
        return best_image, applied_rotation_deg, residual

    @staticmethod
    def _grid_skew_angle_from_centers(centers: np.ndarray) -> float:
        if centers.shape[0] < 20:
            return 0.0
        rect = cv2.minAreaRect(centers.reshape(-1, 1, 2).astype(np.float32))
        rw, rh = rect[1]
        ang = float(rect[2])
        if rw < rh:
            ang += 90.0
        return MRCFinalProcessor._normalize_angle_to_45(ang)

    def _deskew_by_rect_grid(
        self,
        image_bgr: np.ndarray,
        border_value: Tuple[int, int, int],
        min_abs_angle_deg: float = 0.12,
    ) -> Tuple[np.ndarray, float]:
        try:
            mask = self._build_color_invariant_mask(image_bgr)
            rects = self._detect_rectangles(image_bgr, mask)
            if len(rects) < 50:
                return image_bgr, 0.0
            centers = np.array([r["center"] for r in rects], dtype=np.float32)
            ang = self._grid_skew_angle_from_centers(centers)
            if abs(ang) < min_abs_angle_deg:
                return image_bgr, 0.0
            return self._rotate_keep_size(image_bgr, -ang, border_value), -ang
        except Exception:
            return image_bgr, 0.0

    # ------------ rectangle detection + labeling (from GBA1) ------------
    @staticmethod
    def _build_color_invariant_mask(image_bgr: np.ndarray) -> np.ndarray:
        lab = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2LAB)
        a = lab[:, :, 1].astype(np.float32) - 128.0
        b = lab[:, :, 2].astype(np.float32) - 128.0
        chroma = np.sqrt(a * a + b * b)
        chroma = cv2.GaussianBlur(chroma, (3, 3), 0)
        chroma_u8 = (255 * (chroma / (np.max(chroma) + 1e-6))).astype(np.uint8)
        _, mask_chroma = cv2.threshold(chroma_u8, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

        gray = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2GRAY)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        gray_enh = clahe.apply(gray)
        tophat = cv2.morphologyEx(
            gray_enh,
            cv2.MORPH_TOPHAT,
            cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (21, 21)),
        )
        mask_gray = cv2.adaptiveThreshold(
            tophat,
            255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY,
            31,
            -2,
        )

        mask = mask_chroma if int(np.count_nonzero(mask_chroma)) >= 3000 else cv2.bitwise_or(mask_chroma, mask_gray)
        return cv2.morphologyEx(
            mask,
            cv2.MORPH_CLOSE,
            cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3)),
            iterations=1,
        )

    def _detect_rectangles(self, image_bgr: np.ndarray, mask: np.ndarray) -> List[RectInfo]:
        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        rects: List[RectInfo] = []
        for cnt in contours:
            area = float(cv2.contourArea(cnt))
            if area < self.min_area or area > self.max_area:
                continue
            rect = cv2.minAreaRect(cnt)
            rw, rh = rect[1]
            if rw < 1.0 or rh < 1.0:
                continue
            aspect = max(rw, rh) / min(rw, rh)
            if aspect > self.max_aspect_ratio:
                continue
            x, y, w, h = cv2.boundingRect(cnt)
            box = cv2.boxPoints(rect)
            rects.append(
                {
                    "bbox": (int(x), int(y), int(w), int(h)),
                    "center": (float(x + w / 2.0), float(y + h / 2.0)),
                    "area": area,
                    "aspect": float(aspect),
                    "rotated_box": np.int32(np.round(box)).tolist(),
                }
            )

        if len(rects) > 120:
            med_area = float(np.median([r["area"] for r in rects]))
            for r in rects:
                r["quality"] = abs(r["area"] - med_area) / max(med_area, 1e-6) + (r["aspect"] - 1.0) * 0.7
            rects = sorted(rects, key=lambda x: float(x["quality"]))[:120]
        return rects

    def _assign_labels_by_mapping_table(self, rects: List[RectInfo]) -> List[Dict[str, Any]]:
        rects_use = sorted(rects, key=lambda r: float(r.get("area", 0.0)), reverse=True)[:100]
        if len(rects_use) < 100:
            raise RuntimeError(f"可用于编号的矩形不足100个，当前: {len(rects_use)}")

        centers = np.array([r["center"] for r in rects_use], dtype=np.float32)
        xs = np.sort(centers[:, 0])
        ys = np.sort(centers[:, 1])
        x_centers = [float(np.mean(xs[i * 10 : (i + 1) * 10])) for i in range(10)]
        y_centers = [float(np.mean(ys[i * 10 : (i + 1) * 10])) for i in range(10)]

        grid_points: List[Tuple[int, int, float, float]] = []
        for rr in range(10):
            for cc in range(10):
                grid_points.append((rr, cc, x_centers[cc], y_centers[rr]))

        pairs: List[Tuple[float, int, int]] = []
        for i, r in enumerate(rects_use):
            cx, cy = r["center"]
            for j, (_rr, _cc, gx, gy) in enumerate(grid_points):
                d = (float(cx) - gx) ** 2 + (float(cy) - gy) ** 2
                pairs.append((d, i, j))
        pairs.sort(key=lambda x: x[0])

        used_rect = set()
        used_grid = set()
        assign: Dict[int, int] = {}
        for _d, i, j in pairs:
            if i in used_rect or j in used_grid:
                continue
            used_rect.add(i)
            used_grid.add(j)
            assign[i] = j
            if len(assign) == 100:
                break
        if len(assign) < 100:
            raise RuntimeError(f"映射编号唯一匹配失败，已匹配: {len(assign)}")

        mapped: List[Dict[str, Any]] = []
        for i, r in enumerate(rects_use):
            rr, cc, _gx, _gy = grid_points[assign[i]]
            label = self.mapping_grid_10x10[rr][cc]
            mapped.append(
                {
                    "label": label,
                    "group_id": int(label.split("-")[0]),
                    "rect_id": int(label.split("-")[1]),
                    "center": (float(r["center"][0]), float(r["center"][1])),
                    "bbox": tuple(int(v) for v in r["bbox"]),
                    "rotated_box": r["rotated_box"],
                }
            )
        mapped.sort(key=lambda x: (x["group_id"], x["rect_id"]))
        return mapped

    def _detect_rectangles_with_params(
        self,
        image_bgr: np.ndarray,
        mask: np.ndarray,
        min_area: float,
        max_area: float,
        max_aspect_ratio: float,
    ) -> List[RectInfo]:
        contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        rects: List[RectInfo] = []
        for cnt in contours:
            area = float(cv2.contourArea(cnt))
            if area < min_area or area > max_area:
                continue
            rect = cv2.minAreaRect(cnt)
            rw, rh = rect[1]
            if rw < 1.0 or rh < 1.0:
                continue
            aspect = max(rw, rh) / min(rw, rh)
            if aspect > max_aspect_ratio:
                continue
            x, y, w, h = cv2.boundingRect(cnt)
            box = cv2.boxPoints(rect)
            rects.append(
                {
                    "bbox": (int(x), int(y), int(w), int(h)),
                    "center": (float(x + w / 2.0), float(y + h / 2.0)),
                    "area": area,
                    "aspect": float(aspect),
                    "rotated_box": np.int32(np.round(box)).tolist(),
                }
            )
        if len(rects) > 120:
            med_area = float(np.median([r["area"] for r in rects]))
            for r in rects:
                r["quality"] = abs(r["area"] - med_area) / max(med_area, 1e-6) + (r["aspect"] - 1.0) * 0.7
            rects = sorted(rects, key=lambda x: float(x["quality"]))[:120]
        return rects

    def _assign_labels_relaxed_for_corner_orient(self, rects: List[RectInfo]) -> List[Dict[str, Any]]:
        """与 lingxing 一致：不足 100 个矩形时仍可部分匹配，用于四角波峰波谷定向。"""
        rects_use = sorted(rects, key=lambda r: float(r.get("area", 0.0)), reverse=True)[:100]
        if len(rects_use) < 40:
            raise RuntimeError(f"可用于定向的矩形过少，当前: {len(rects_use)}")
        centers = np.array([r["center"] for r in rects_use], dtype=np.float32)
        xs = np.sort(centers[:, 0])
        ys = np.sort(centers[:, 1])
        x_centers = [float(np.mean(xs[i * 10 : (i + 1) * 10])) for i in range(10)]
        y_centers = [float(np.mean(ys[i * 10 : (i + 1) * 10])) for i in range(10)]

        grid_points: List[Tuple[int, int, float, float]] = []
        for rr in range(10):
            for cc in range(10):
                grid_points.append((rr, cc, x_centers[cc], y_centers[rr]))
        pairs: List[Tuple[float, int, int]] = []
        for i, r in enumerate(rects_use):
            cx, cy = r["center"]
            for j, (_rr, _cc, gx, gy) in enumerate(grid_points):
                d = (float(cx) - gx) ** 2 + (float(cy) - gy) ** 2
                pairs.append((d, i, j))
        pairs.sort(key=lambda x: x[0])
        used_rect: set = set()
        used_grid: set = set()
        assign: Dict[int, int] = {}
        for _d, i, j in pairs:
            if i in used_rect or j in used_grid:
                continue
            used_rect.add(i)
            used_grid.add(j)
            assign[i] = j
            if len(assign) == len(rects_use):
                break
        if len(assign) < max(30, len(rects_use) // 2):
            raise RuntimeError(f"定向用映射匹配不足，已匹配: {len(assign)}")
        mapped: List[Dict[str, Any]] = []
        for i, r in enumerate(rects_use):
            if i not in assign:
                continue
            rr, cc, _gx, _gy = grid_points[assign[i]]
            label = self.mapping_grid_10x10[rr][cc]
            mapped.append(
                {
                    "label": label,
                    "group_id": int(label.split("-")[0]),
                    "rect_id": int(label.split("-")[1]),
                    "center": (float(r["center"][0]), float(r["center"][1])),
                    "bbox": tuple(int(v) for v in r["bbox"]),
                    "rotated_box": r["rotated_box"],
                }
            )
        mapped.sort(key=lambda x: (x["group_id"], x["rect_id"]))
        return mapped

    def _compute_stripe_mismatch_score(
        self,
        work_bgr: np.ndarray,
        profile_roi_pad: int = 2,
    ) -> Tuple[int, List[Dict[str, Any]]]:
        expected = self._default_expected_stripe_counts()
        mask = self._build_color_invariant_mask(work_bgr)
        rects = self._detect_rectangles(work_bgr, mask)
        mapped = self._assign_labels_by_mapping_table(rects)
        h, w = work_bgr.shape[:2]
        mismatch = 0
        for m in mapped:
            if int(m.get("rect_id", -1)) != 1:
                continue
            gid = int(m["group_id"])
            x1, y1, x2, y2 = self._profile_roi_xyxy_right_pad_only(m["bbox"], profile_roi_pad, w, h)
            metrics = self._extract_profile_metrics(work_bgr[y1:y2, x1:x2])
            stripe_n = int(len(metrics["peaks"]) + len(metrics["valleys"]))
            mismatch += abs(stripe_n - int(expected.get(gid, stripe_n)))
        return int(mismatch), mapped

    def _apply_orientation_by_stripe_consistency(
        self,
        aligned_bgr: np.ndarray,
        border_value: Tuple[int, int, int],
        profile_roi_pad: int = 2,
    ) -> Tuple[np.ndarray, int, Dict[str, int], np.ndarray]:
        """在 0/90/180/270 中选一组，使各组 rect1 条纹数最接近预期表（替代四角波峰波谷启发式）。"""
        best_mismatch = 10**9
        best_image = aligned_bgr
        best_rot = 0
        best_counts: Dict[str, int] = {"TL": 0, "TR": 0, "BL": 0, "BR": 0}
        best_debug = aligned_bgr.copy()

        for rot_i in range(4):
            work = aligned_bgr
            if rot_i > 0:
                work = self._rotate_keep_size_with_pad(aligned_bgr, 90.0 * rot_i, border_value)
            try:
                mismatch, mapped = self._compute_stripe_mismatch_score(work, profile_roi_pad=profile_roi_pad)
            except Exception:
                continue
            if mismatch >= best_mismatch:
                continue
            best_mismatch = mismatch
            best_rot = rot_i
            best_image = work
            h, w = work.shape[:2]
            corners = {
                "TL": (0.0, 0.0),
                "TR": (float(w), 0.0),
                "BL": (0.0, float(h)),
                "BR": (float(w), float(h)),
            }
            rect1 = [m for m in mapped if int(m.get("rect_id", -1)) == 1]
            if len(rect1) < 4:
                rect1 = sorted(mapped, key=lambda m: int(m.get("group_id", 9999)))[: min(12, len(mapped))]
            debug = work.copy()
            counts: Dict[str, int] = {}
            for name, cxy in corners.items():
                if not rect1:
                    counts[name] = 0
                    continue
                picked = min(
                    rect1,
                    key=lambda r, cxy=cxy: float((r["center"][0] - cxy[0]) ** 2 + (r["center"][1] - cxy[1]) ** 2),
                )
                x1, y1, x2, y2 = self._profile_roi_xyxy_right_pad_only(picked["bbox"], profile_roi_pad, w, h)
                metrics = self._extract_profile_metrics(work[y1:y2, x1:x2])
                stripe_n = int(len(metrics["peaks"]) + len(metrics["valleys"]))
                counts[name] = stripe_n
                color = (0, 255, 0) if name == "TL" else (255, 0, 0)
                cv2.rectangle(debug, (x1, y1), (x2, y2), color, 2)
                cv2.putText(
                    debug,
                    f"{name}:{stripe_n}",
                    (x1, max(12, y1 - 4)),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.5,
                    color,
                    1,
                    cv2.LINE_AA,
                )
            best_counts = counts
            best_debug = debug

        if best_mismatch >= 10**9:
            raise RuntimeError("条纹一致性定向失败：四个方向均无法完成编号。")
        return best_image, best_rot, best_counts, best_debug

    def _apply_corner_min_stripe_orientation(
        self,
        aligned_bgr: np.ndarray,
        border_value: Tuple[int, int, int],
        profile_roi_pad: int = 2,
    ) -> Tuple[np.ndarray, int, Dict[str, int], np.ndarray]:
        """摆正后：比较四角「第一组矩形 rect_id=1」剖面波峰+波谷数，直到左上角 TL 为最少；否则每次顺时针 90°。"""
        final = aligned_bgr.copy()
        best_counts: Dict[str, int] = {"TL": 9999, "TR": 9999, "BL": 9999, "BR": 9999}
        best_debug = aligned_bgr.copy()
        for rot_i in range(4):
            rect_mask = self._build_color_invariant_mask(final)
            rects = self._detect_rectangles_with_params(
                final,
                rect_mask,
                self._corner_orient_min_area,
                self._corner_orient_max_area,
                self._corner_orient_max_aspect_ratio,
            )
            mapped = self._assign_labels_relaxed_for_corner_orient(rects)
            rect1 = [m for m in mapped if int(m.get("rect_id", -1)) == 1]
            if len(rect1) < 4:
                rect1 = sorted(mapped, key=lambda m: int(m.get("group_id", 9999)))[: min(12, len(mapped))]
                if len(rect1) < 4:
                    raise RuntimeError("候选矩形不足，无法完成四角波峰波谷定向。")
            h, w = final.shape[:2]
            corners = {"TL": (0.0, 0.0), "TR": (float(w), 0.0), "BL": (0.0, float(h)), "BR": (float(w), float(h))}
            picked = {
                name: min(rect1, key=lambda r, cxy=cxy: float((r["center"][0] - cxy[0]) ** 2 + (r["center"][1] - cxy[1]) ** 2))
                for name, cxy in corners.items()
            }
            debug = final.copy()
            counts: Dict[str, int] = {}
            for name, m in picked.items():
                x1, y1, x2, y2 = self._profile_roi_xyxy_right_pad_only(m["bbox"], profile_roi_pad, w, h)
                metrics = self._extract_profile_metrics(final[y1:y2, x1:x2])
                stripe_n = int(len(metrics["peaks"]) + len(metrics["valleys"]))
                counts[name] = stripe_n
                color = (0, 255, 0) if name == "TL" else (255, 0, 0)
                cv2.rectangle(debug, (x1, y1), (x2, y2), color, 2)
                cv2.putText(
                    debug,
                    f"{name}:{stripe_n}",
                    (x1, max(12, y1 - 4)),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.5,
                    color,
                    1,
                    cv2.LINE_AA,
                )
            best_counts = counts
            best_debug = debug
            if min(counts, key=counts.get) == "TL":
                return final, rot_i, best_counts, best_debug
            final = self._rotate_keep_size_with_pad(final, 90.0, border_value)
        return final, 3, best_counts, best_debug

    @staticmethod
    def _draw_results_mapped(image_bgr: np.ndarray, mapped_rects: List[Dict[str, Any]]) -> np.ndarray:
        out = image_bgr.copy()
        for m in mapped_rects:
            box = np.array(m["rotated_box"], dtype=np.int32).reshape((-1, 1, 2))
            cv2.polylines(out, [box], isClosed=True, color=(0, 0, 255), thickness=2)
            box_pts = np.array(m["rotated_box"], dtype=np.int32)
            x = int(np.min(box_pts[:, 0]))
            y = int(np.min(box_pts[:, 1]))
            cv2.putText(
                out,
                m["label"],
                (x, max(18, y - 4)),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.45,
                (0, 0, 255),
                1,
                cv2.LINE_AA,
            )
        return out

    @staticmethod
    def _build_overview_panel(original: np.ndarray, aligned: np.ndarray, mask: np.ndarray, final: np.ndarray) -> np.ndarray:
        h, w = original.shape[:2]
        half_w, half_h = w // 2, h // 2
        p1 = cv2.resize(original, (half_w, half_h), interpolation=cv2.INTER_AREA)
        p2 = cv2.resize(aligned, (half_w, half_h), interpolation=cv2.INTER_AREA)
        p3 = cv2.cvtColor(mask, cv2.COLOR_GRAY2BGR)
        p3 = cv2.resize(p3, (half_w, half_h), interpolation=cv2.INTER_NEAREST)
        p4 = cv2.resize(final, (half_w, half_h), interpolation=cv2.INTER_AREA)
        cv2.putText(p1, "Original", (12, 28), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2, cv2.LINE_AA)
        cv2.putText(p2, "Aligned", (12, 28), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2, cv2.LINE_AA)
        cv2.putText(p3, "Rectangle Mask", (12, 28), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2, cv2.LINE_AA)
        cv2.putText(p4, "Labeled 100 Rectangles", (12, 28), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2, cv2.LINE_AA)
        return np.vstack([np.hstack([p1, p2]), np.hstack([p3, p4])])

    @staticmethod
    def _extract_profile_metrics(roi_color: np.ndarray) -> Dict[str, Any]:
        # 与 MRC_bofeng.extract_profile_metrics 保持一致（波峰波谷算法不改）
        if roi_color.size == 0 or roi_color.shape[1] < 8:
            return {
                "line_raw": np.array([], dtype=np.float32),
                "row_idx": -1,
                "peaks": [],
                "valleys": [],
                "pair_n": 0,
                "c_mean": 0.0,
                "peak_vals_raw": [],
                "valley_vals_raw": [],
            }
        roi_gray = cv2.cvtColor(roi_color, cv2.COLOR_BGR2GRAY).astype(np.float32)
        row_idx = int(np.argmax(np.std(roi_gray, axis=1)))
        line_raw = roi_gray[row_idx, :]
        d = np.diff(line_raw)
        peaks: List[int] = []
        valleys: List[int] = []
        for i in range(1, len(d)):
            if d[i - 1] > 0 and d[i] <= 0:
                peaks.append(i)
            if d[i - 1] < 0 and d[i] >= 0:
                valleys.append(i)
        peak_vals_raw = [float(line_raw[i]) for i in peaks]
        valley_vals_raw = [float(line_raw[i]) for i in valleys]
        pair_n = min(len(peak_vals_raw), len(valley_vals_raw))
        c_vals: List[float] = []
        for i in range(pair_n):
            p, v = peak_vals_raw[i], valley_vals_raw[i]
            den = p + v
            if den > 1e-6 and p > v:
                c_vals.append((p - v) / den)
        c_mean = float(np.mean(np.array(c_vals, dtype=np.float32))) if c_vals else 0.0
        return {
            "line_raw": line_raw,
            "row_idx": row_idx,
            "peaks": peaks,
            "valleys": valleys,
            "pair_n": pair_n,
            "c_mean": c_mean,
            "peak_vals_raw": peak_vals_raw,
            "valley_vals_raw": valley_vals_raw,
        }

    @staticmethod
    def _save_group_results_excel(rows: List[Dict[str, Any]], excel_path: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "group_rect1_results"
        ws.append(
            [
                "group_id",
                "label",
                "peak_gray_values",
                "valley_gray_values",
                "valid_pair_count",
                "stripe_count_n",
                "expected_stripe_count_n",
                "is_abnormal",
                "result_c_mean",
            ]
        )
        for row in rows:
            ws.append(
                [
                    int(row["group_id"]),
                    str(row["label"]),
                    ",".join(f"{float(v):.3f}" for v in row["peak_vals_raw"]),
                    ",".join(f"{float(v):.3f}" for v in row["valley_vals_raw"]),
                    int(row["pair_n"]),
                    int(row["stripe_n"]),
                    int(row["expected_stripe_n"]),
                    int(row["is_abnormal"]),
                    float(row["c_mean"]),
                ]
            )
        for col_cells in ws.columns:
            max_len = 0
            letter = col_cells[0].column_letter
            for cell in col_cells:
                max_len = max(max_len, len("" if cell.value is None else str(cell.value)))
            ws.column_dimensions[letter].width = min(max(12, max_len + 2), 80)
        wb.save(excel_path)

    @staticmethod
    def _save_group_curve_plot(rows: List[Dict[str, Any]], target: float, plot_path: str) -> None:
        if not rows:
            raise RuntimeError("无可用于绘图的组数据。")
        normal_rows = [r for r in rows if int(r.get("is_abnormal", 0)) == 0]
        abnormal_rows = [r for r in rows if int(r.get("is_abnormal", 0)) == 1]
        group_ids = [int(r["group_id"]) for r in rows]
        c_means = [float(r["c_mean"]) for r in rows]
        fig, ax = plt.subplots(figsize=(11, 5))
        ax.plot(group_ids, c_means, lw=1.2, color="#1f77b4", alpha=0.9, label="Rect1 result")
        if normal_rows:
            ax.scatter(
                [int(r["group_id"]) for r in normal_rows],
                [float(r["c_mean"]) for r in normal_rows],
                s=42,
                c="#1f77b4",
                marker="o",
                edgecolors="#1f77b4",
                linewidths=1.0,
                zorder=3,
                label="Normal",
            )
        if abnormal_rows:
            ax.scatter(
                [int(r["group_id"]) for r in abnormal_rows],
                [float(r["c_mean"]) for r in abnormal_rows],
                s=62,
                facecolors="none",
                edgecolors="#d62728",
                marker="o",
                linewidths=2.0,
                zorder=4,
                label="Abnormal (hollow)",
            )
        ax.axhline(target, color="#ff0000", lw=2.8, ls="--", label=f"Standard={target:.3f}")
        ax.fill_between(group_ids, target - 0.005, target + 0.005, color="#ff0000", alpha=0.10)
        ax.text(
            group_ids[-1],
            target + 0.006,
            f"  Standard line: {target:.3f}",
            color="#ff0000",
            fontsize=10,
            fontweight="bold",
            ha="right",
            va="bottom",
        )
        ax.set_xlabel("Group ID")
        ax.set_ylabel("Result value (C_mean)")
        ax.set_title("Group Rect1 Result Curve (with 0.03 standard)")
        ax.grid(True, alpha=0.3)
        ax.set_xticks(group_ids)
        ax.legend(loc="best", fontsize=9)
        plt.tight_layout()
        plt.savefig(plot_path, dpi=150, bbox_inches="tight")
        plt.close(fig)

    @staticmethod
    def _default_expected_stripe_counts() -> Dict[int, int]:
        # 用户给定“每组所含明暗线条总数 n”标准表
        return {
            1: 7, 2: 7, 3: 7, 4: 7,
            5: 9, 6: 9, 7: 9,
            8: 11, 9: 11, 10: 11,
            11: 13, 12: 13,
            13: 15, 14: 15, 15: 15,
            16: 17,
            17: 11,
            18: 13, 19: 13, 20: 13,
            21: 15, 22: 15, 23: 17,
            24: 17,
            25: 19,
        }

    @staticmethod
    def _parse_expected_pairs(spec: str) -> Dict[int, int]:
        expected: Dict[int, int] = {}
        if not spec.strip():
            return expected
        for token in spec.split(","):
            item = token.strip()
            if not item:
                continue
            if ":" not in item:
                raise ValueError(f"expected-pairs格式错误: {item}")
            k, v = item.split(":", 1)
            gid = int(k.strip())
            cnt = int(v.strip())
            expected[gid] = cnt
        return expected

    @staticmethod
    def _infer_expected_pairs(rows: List[Dict[str, Any]], window_radius: int = 2) -> Dict[int, int]:
        if not rows:
            return {}
        rows_sorted = sorted(rows, key=lambda x: int(x["group_id"]))
        pair_vals = [int(r["stripe_n"]) for r in rows_sorted]
        gids = [int(r["group_id"]) for r in rows_sorted]
        inferred: Dict[int, int] = {}
        n = len(rows_sorted)
        for i, gid in enumerate(gids):
            l = max(0, i - window_radius)
            rr = min(n, i + window_radius + 1)
            local = pair_vals[l:rr]
            inferred[gid] = int(round(float(np.median(np.array(local, dtype=np.float32)))))
        return inferred

    @staticmethod
    def _profile_roi_xyxy_right_pad_only(
        bbox: Tuple[int, int, int, int],
        right_pad: int,
        img_w: int,
        img_h: int,
    ) -> Tuple[int, int, int, int]:
        """与 MRC_bofeng.profile_roi_xyxy_right_pad_only 一致：左/上/下为 bbox；仅向右多取列。"""
        x, y, w, h = (int(v) for v in bbox)
        pad = max(0, int(right_pad))
        x1, y1 = max(0, x), max(0, y)
        x2 = min(img_w, x + w + pad)
        y2 = min(img_h, y + h)
        return x1, y1, x2, y2

    @staticmethod
    def _attach_abnormal_flags(
        rows: List[Dict[str, Any]],
        expected_pairs: Dict[int, int],
        pair_tolerance: int,
    ) -> None:
        for row in rows:
            gid = int(row["group_id"])
            expected = int(expected_pairs.get(gid, int(row["stripe_n"])))
            stripe_n = int(row["stripe_n"])
            row["expected_stripe_n"] = expected
            row["is_abnormal"] = int(abs(stripe_n - expected) > max(0, int(pair_tolerance)))

    @staticmethod
    def _pick_min_resolvable_group(rows: List[Dict[str, Any]], threshold: float) -> Tuple[Optional[int], Optional[float]]:
        """最小分辨组定义：满足 c_mean > threshold 且 is_abnormal == 0 的最高 group_id。"""
        valid = [
            r for r in rows
            if int(r.get("is_abnormal", 1)) == 0 and float(r.get("c_mean", 0.0)) > float(threshold)
        ]
        if not valid:
            return None, None
        best = max(valid, key=lambda x: int(x.get("group_id", -1)))
        return int(best["group_id"]), float(best["c_mean"])

    def process_image(
        self,
        image_path: str,
        output_dir: str,
        target_value: float = 0.03,
        expected_pairs_spec: str = "",
        pair_tolerance: int = 0,
        profile_roi_pad: int = 2,
        enable_corner_orientation: bool = True,
    ) -> Dict[str, Any]:
        image = self._imread_unicode(image_path)
        if image is None:
            raise FileNotFoundError(f"无法读取图像: {image_path}")

        angle_deg, _, _, _ = self._measure_side_angle(image)
        border_value = self._estimate_dark_border_value(image)
        aligned, applied_rotation_deg, residual_angle = self._refine_rotation_with_verification(
            image, angle_deg, border_value
        )
        grid_aligned, grid_correction = self._deskew_by_rect_grid(aligned, border_value)
        applied_rotation_deg += grid_correction
        if abs(grid_correction) > 1e-6:
            try:
                residual_angle, _, _, _ = self._measure_side_angle(grid_aligned)
            except Exception:
                pass
            else:
                aligned = grid_aligned
        work_bgr = aligned
        corner_rot_90_count = 0
        corner_counts: Dict[str, int] = {}
        if enable_corner_orientation:
            work_bgr, corner_rot_90_count, corner_counts, _corner_debug = self._apply_orientation_by_stripe_consistency(
                aligned, border_value, profile_roi_pad=profile_roi_pad
            )

        rect_mask = self._build_color_invariant_mask(work_bgr)
        rect_candidates = self._detect_rectangles(work_bgr, rect_mask)
        mapped_rects = self._assign_labels_by_mapping_table(rect_candidates)
        final_img = self._draw_results_mapped(work_bgr, mapped_rects)
        overview = self._build_overview_panel(image, aligned, rect_mask, final_img)

        os.makedirs(output_dir, exist_ok=True)
        stem = os.path.splitext(os.path.basename(image_path))[0]
        ext = os.path.splitext(image_path)[1] or ".png"
        out_aligned = os.path.join(output_dir, f"{stem}_a{ext}")
        out_output = os.path.join(output_dir, f"{stem}_labels{ext}")
        out_overview = os.path.join(output_dir, f"{stem}_ov{ext}")
        out_group_excel = os.path.join(output_dir, f"{stem}_res.xlsx")
        out_group_curve = os.path.join(output_dir, f"{stem}_curve.png")
        out_corner_debug = os.path.join(output_dir, f"{stem}_corner_debug{ext}")

        for p, img in [
            (out_aligned, aligned),
            (out_output, final_img),
            (out_overview, overview),
        ]:
            if not self._imwrite_unicode(p, img):
                raise IOError(f"保存失败: {p}")
        if enable_corner_orientation:
            if not self._imwrite_unicode(out_corner_debug, _corner_debug):
                raise IOError(f"保存失败: {out_corner_debug}")

        group_rect1_rows: List[Dict[str, Any]] = []
        rect1_all = [m for m in mapped_rects if int(m["rect_id"]) == 1]
        rect1_all.sort(key=lambda x: int(x["group_id"]))
        ih, iw = int(work_bgr.shape[0]), int(work_bgr.shape[1])
        for m in rect1_all:
            x1, y1, x2, y2 = self._profile_roi_xyxy_right_pad_only(m["bbox"], profile_roi_pad, iw, ih)
            roi = work_bgr[y1:y2, x1:x2]
            metrics = self._extract_profile_metrics(roi)
            group_rect1_rows.append(
                {
                    "group_id": int(m["group_id"]),
                    "label": str(m["label"]),
                    "peak_vals_raw": list(metrics["peak_vals_raw"]),
                    "valley_vals_raw": list(metrics["valley_vals_raw"]),
                    "pair_n": int(metrics["pair_n"]),
                    "stripe_n": int(len(metrics["peaks"]) + len(metrics["valleys"])),
                    "c_mean": float(metrics["c_mean"]),
                }
            )
        expected_pairs = self._parse_expected_pairs(expected_pairs_spec)
        if not expected_pairs:
            expected_pairs = self._default_expected_stripe_counts()
        if not expected_pairs:
            expected_pairs = self._infer_expected_pairs(group_rect1_rows)
        self._attach_abnormal_flags(group_rect1_rows, expected_pairs=expected_pairs, pair_tolerance=pair_tolerance)
        self._save_group_results_excel(group_rect1_rows, out_group_excel)
        self._save_group_curve_plot(group_rect1_rows, target_value, out_group_curve)
        min_group_id, min_group_c_mean = self._pick_min_resolvable_group(group_rect1_rows, target_value)

        out_summary_json = os.path.join(output_dir, f"{stem}_summary.json")
        summary = {
            "image_name": os.path.basename(image_path),
            "min_resolvable_group_id": min_group_id,
            "min_resolvable_c_mean": min_group_c_mean,
            "threshold": float(target_value),
            "valid_rule": "c_mean > threshold and is_abnormal == 0",
        }
        with open(out_summary_json, "w", encoding="utf-8") as f:
            json.dump(summary, f, ensure_ascii=False, indent=2)

        out: Dict[str, Any] = {
            "candidate_count": len(rect_candidates),
            "mapped_count": len(mapped_rects),
            "side_angle_deg": angle_deg,
            "applied_rotation_deg": applied_rotation_deg,
            "residual_angle_deg": residual_angle,
            "corner_orientation_enabled": bool(enable_corner_orientation),
            "corner_rot_90_count": int(corner_rot_90_count),
            "corner_stripe_counts": dict(corner_counts) if corner_counts else {},
            "mapped_rects": mapped_rects,
            "aligned_path": out_aligned,
            "output_path": out_output,
            "overview_path": out_overview,
            "group_excel_path": out_group_excel,
            "group_curve_path": out_group_curve,
            "summary_json_path": out_summary_json,
            "min_resolvable_group_id": min_group_id,
            "min_resolvable_c_mean": min_group_c_mean,
        }
        if enable_corner_orientation:
            out["corner_debug_path"] = out_corner_debug
        return out


def collect_images(folder: str) -> List[str]:
    exts = ["*.bmp", "*.jpg", "*.jpeg", "*.png", "*.tif", "*.tiff"]
    paths: List[str] = []
    for ext in exts:
        paths.extend(glob.glob(os.path.join(folder, ext)))
    filtered = []
    derived_suffixes = ("_a", "_labels", "_ov", "_curve", "_corner_debug")
    for p in paths:
        stem = os.path.splitext(os.path.basename(p))[0].lower()
        if stem.endswith(derived_suffixes):
            continue
        filtered.append(p)
    return sorted(list(set(filtered)))


def pick_batch_min_resolvable_group(group_ids: List[int]) -> Optional[int]:
    """批量判定：出现次数最多的组号；并列时取较小组号（向下兼容）。"""
    if not group_ids:
        return None
    counts: Dict[int, int] = {}
    for gid in group_ids:
        counts[gid] = counts.get(gid, 0) + 1
    max_count = max(counts.values())
    tied = [gid for gid, cnt in counts.items() if cnt == max_count]
    return min(tied)


def _nice_y_tick_step(max_count: int) -> int:
    """根据最大频次自动选择整数纵轴刻度步长（1/2/5/10/20…）。"""
    if max_count <= 0:
        return 1
    if max_count <= 5:
        return 1
    if max_count <= 12:
        return 2
    if max_count <= 30:
        return 5
    if max_count <= 60:
        return 10
    if max_count <= 150:
        return 20
    if max_count <= 300:
        return 50

    target_ticks = 6
    raw = max_count / target_ticks
    exp = 10 ** math.floor(math.log10(raw))
    fraction = raw / exp
    if fraction <= 1:
        nice = 1
    elif fraction <= 2:
        nice = 2
    elif fraction <= 5:
        nice = 5
    else:
        nice = 10
    return max(1, int(nice * exp))


def _build_integer_y_axis(max_count: int) -> Tuple[int, List[int]]:
    """构建覆盖峰值且刻度均为整数的纵轴范围。"""
    step = _nice_y_tick_step(max_count)
    y_top = int(math.ceil(max_count / step) * step)
    if y_top <= max_count:
        y_top += step
    ticks = list(range(0, y_top + 1, step))
    return y_top, ticks


def _smooth_distribution_curve(
    groups: List[int],
    values: List[float],
    num_points: int = 400,
) -> Tuple[np.ndarray, np.ndarray]:
    """在离散组号之间插值，生成平滑的频率分布曲线。"""
    x_arr = np.asarray(groups, dtype=float)
    y_arr = np.asarray(values, dtype=float)

    if len(x_arr) == 1:
        pad = 0.6
        x_smooth = np.linspace(x_arr[0] - pad, x_arr[0] + pad, num_points)
        y_smooth = np.full_like(x_smooth, y_arr[0], dtype=float)
        return x_smooth, y_smooth

    x_min, x_max = float(x_arr.min()), float(x_arr.max())
    span = max(x_max - x_min, 1.0)
    x_pad_min = x_min - span * 0.04
    x_pad_max = x_max + span * 0.04
    x_dense = np.linspace(x_pad_min, x_pad_max, num_points)

    try:
        from scipy.interpolate import PchipInterpolator

        interpolator = PchipInterpolator(x_arr, y_arr, extrapolate=False)
        y_dense = interpolator(x_dense)
        y_dense = np.clip(y_dense, 0.0, None)
        y_dense = np.nan_to_num(y_dense, nan=0.0)
    except Exception:
        y_dense = np.interp(x_dense, x_arr, y_arr)
        y_dense = np.clip(y_dense, 0.0, None)

    return x_dense, y_dense


def _ensure_plot_fonts() -> None:
    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "SimSun", "Noto Sans CJK SC", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False


def save_batch_group_distribution_plot(
    group_ids: List[int],
    plot_path: str,
    batch_group_id: Optional[int] = None,
) -> Optional[int]:
    """绘制批量最小可分辨组号频率分布曲线图。"""
    if not group_ids:
        raise RuntimeError("没有可用于绘制分布图的有效组号。")

    _ensure_plot_fonts()

    counts: Dict[int, int] = {}
    for gid in group_ids:
        counts[gid] = counts.get(gid, 0) + 1

    if batch_group_id is None:
        batch_group_id = pick_batch_min_resolvable_group(group_ids)

    groups = sorted(counts.keys())
    values = [float(counts[g]) for g in groups]
    total = len(group_ids)
    max_count = int(max(values))
    y_top, y_ticks = _build_integer_y_axis(max_count)

    x_smooth, y_smooth = _smooth_distribution_curve(groups, values)

    fig, ax = plt.subplots(figsize=(10.5, 5.8))
    fig.patch.set_facecolor("#fafafa")
    ax.set_facecolor("#ffffff")

    ax.fill_between(x_smooth, y_smooth, 0, color="#4a90d9", alpha=0.22, zorder=1)
    ax.plot(
        x_smooth,
        y_smooth,
        color="#1f5a99",
        lw=2.4,
        alpha=0.95,
        zorder=2,
        label="频率分布曲线",
    )
    ax.scatter(
        groups,
        values,
        s=72,
        c="#ffffff",
        edgecolors="#1f5a99",
        linewidths=2.0,
        zorder=4,
        label="实测频数",
    )

    peak_group = batch_group_id if batch_group_id is not None else groups[int(np.argmax(values))]
    peak_count = counts.get(peak_group, max_count)

    if batch_group_id is not None and batch_group_id in counts:
        ax.axvline(
            batch_group_id,
            color="#c0392b",
            lw=1.8,
            ls="--",
            alpha=0.85,
            zorder=3,
            label=f"判定组号 {batch_group_id}",
        )
        ax.scatter(
            [batch_group_id],
            [peak_count],
            s=120,
            c="#e74c3c",
            edgecolors="#922b21",
            linewidths=1.6,
            zorder=5,
        )
        ax.annotate(
            f"判定组 {batch_group_id}\n{peak_count}/{total} 张",
            xy=(batch_group_id, peak_count),
            xytext=(18, 22),
            textcoords="offset points",
            fontsize=10,
            color="#922b21",
            fontweight="bold",
            bbox=dict(boxstyle="round,pad=0.35", fc="#fff8f0", ec="#e0c4b0", alpha=0.95),
            arrowprops=dict(arrowstyle="->", color="#c0392b", lw=1.3),
        )

    ax.set_xlabel("最小可分辨组号", fontsize=11)
    ax.set_ylabel("图像数量（张）", fontsize=11)
    ax.set_title(
        f"批量最小可分辨组号频率分布（有效图像 {total} 张）",
        fontsize=12,
        fontweight="bold",
        pad=12,
    )

    ax.set_ylim(0, y_top)
    ax.set_yticks(y_ticks)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f"{int(round(v))}"))

    if len(groups) > 1:
        span = groups[-1] - groups[0]
        pad = max(0.5, span * 0.06)
        ax.set_xlim(groups[0] - pad, groups[-1] + pad)
    else:
        ax.set_xlim(groups[0] - 0.8, groups[0] + 0.8)

    ax.set_xticks(groups)
    ax.grid(True, which="major", axis="both", alpha=0.28, linestyle="-", linewidth=0.6)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.legend(loc="upper right", fontsize=9, framealpha=0.92)

    plt.tight_layout()
    plt.savefig(plot_path, dpi=160, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    return batch_group_id


def run_batch_distribution_mode(group_ids: List[int], output_path: str) -> Dict[str, Any]:
    batch_group_id = save_batch_group_distribution_plot(group_ids, output_path)
    counts: Dict[int, int] = {}
    for gid in group_ids:
        counts[gid] = counts.get(gid, 0) + 1
    return {
        "batch_min_resolvable_group_id": batch_group_id,
        "valid_image_count": len(group_ids),
        "count_by_group_id": counts,
        "plot_path": output_path,
    }


if __name__ == "__main__":
    _ensure_plot_fonts()
    _ = font_manager.findfont("DejaVu Sans")

    parser = argparse.ArgumentParser(description="MRC final: rotate-detect + align + rectangle mapping")
    parser.add_argument("--input", default="", help="输入图像路径（可选）")
    parser.add_argument(
        "--output",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "result"),
        help="输出目录（默认：脚本目录/result）",
    )
    parser.add_argument(
        "--mapping",
        default=r"F:\研究生\项目\MRC_progress\映射表.xlsx",
        help="映射表路径（10x10）",
    )
    parser.add_argument("--target", type=float, default=0.03, help="结果曲线标准线，默认0.03")
    parser.add_argument(
        "--expected-pairs",
        default="",
        help="每组预期明暗条数，如 1:3,2:3,3:3；为空时自动按邻域趋势估计",
    )
    parser.add_argument("--pair-tol", type=int, default=0, help="条数容差，默认0（不允许偏差）")
    parser.add_argument(
        "--profile-roi-pad",
        type=int,
        default=2,
        help="1号矩形剖面 ROI：仅向右多取列数，默认2；与 MRC_bofeng 一致",
    )
    parser.add_argument(
        "--no-corner-orient",
        action="store_true",
        help="关闭四角波峰波谷定向（仅保留倾角摆正后的结果）",
    )
    parser.add_argument(
        "--batch-distribution",
        action="store_true",
        help="批量模式：根据多张图像的最小可分辨组号绘制分布图",
    )
    parser.add_argument(
        "--group-ids",
        default="",
        help="批量分布模式下的组号列表，逗号分隔，如 20,21,21,21,22",
    )
    parser.add_argument(
        "--plot-output",
        default="",
        help="批量分布图输出路径（默认：<output>/min_group_distribution.png）",
    )
    args = parser.parse_args()

    if args.batch_distribution:
        raw = [x.strip() for x in str(args.group_ids).split(",") if x.strip()]
        if not raw:
            raise SystemExit("批量分布模式需要 --group-ids 参数。")
        try:
            group_ids = [int(x) for x in raw]
        except ValueError as exc:
            raise SystemExit(f"--group-ids 格式错误: {exc}") from exc

        output_dir = os.path.abspath(args.output)
        os.makedirs(output_dir, exist_ok=True)
        plot_path = args.plot_output.strip() or os.path.join(output_dir, "min_group_distribution.png")
        plot_path = os.path.abspath(plot_path)
        os.makedirs(os.path.dirname(plot_path), exist_ok=True)

        result = run_batch_distribution_mode(group_ids, plot_path)
        print(json.dumps(result, ensure_ascii=False))
        raise SystemExit(0)

    output_dir = os.path.abspath(args.output)
    os.makedirs(output_dir, exist_ok=True)

    if args.input:
        image_paths = [args.input]
    else:
        zheng_candidates = [
            r"F:\研究生\项目\MRC_progress\mrc_zheng.png",
            r"F:\研究生\项目\MRC_progress\mrc_zheng.jpg",
            r"F:\研究生\项目\MRC_progress\mrc_zheng.bmp",
            r"F:\研究生\项目\MRC_progress\mrc_zheng.jpeg",
        ]
        fallback = [
            r"F:\研究生\项目\MRC_progress\mrc_test.png",
            r"F:\研究生\项目\MRC_progress\mrc_daozhi.png",
        ]
        image_paths: List[str] = []
        for p in zheng_candidates:
            if os.path.exists(p):
                image_paths = [p]
                break
        if not image_paths:
            for p in fallback:
                if os.path.exists(p):
                    image_paths = [p]
                    break
        if not image_paths:
            image_paths = collect_images(os.path.dirname(os.path.abspath(__file__)))
        if not image_paths:
            raise FileNotFoundError("未找到可处理图像，请使用 --input 指定。")

    processor = MRCFinalProcessor(mapping_table_path=args.mapping)
    for img_path in image_paths:
        name = os.path.basename(img_path)
        try:
            processor.process_image(
                img_path,
                output_dir=output_dir,
                target_value=args.target,
                expected_pairs_spec=args.expected_pairs,
                pair_tolerance=args.pair_tol,
                profile_roi_pad=args.profile_roi_pad,
                enable_corner_orientation=not args.no_corner_orient,
            )
            print(f"[OK] {name}")
        except Exception as e:
            print(f"[FAIL] {name}: {e}")
