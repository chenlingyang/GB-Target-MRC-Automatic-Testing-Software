#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import argparse
import os
from typing import Any, Dict, List, Optional, Tuple

import cv2
import numpy as np
import openpyxl

RectInfo = Dict[str, Any]


def imread_unicode(path: str) -> Optional[np.ndarray]:
    data = np.fromfile(path, dtype=np.uint8)
    if data.size == 0:
        return None
    return cv2.imdecode(data, cv2.IMREAD_COLOR)


def imwrite_unicode(path: str, image: np.ndarray) -> bool:
    ext = os.path.splitext(path)[1] or ".png"
    ok, buf = cv2.imencode(ext, image)
    if not ok:
        return False
    buf.tofile(path)
    return True


def load_mapping_grid_10x10(mapping_path: str) -> List[List[str]]:
    if not os.path.exists(mapping_path):
        raise FileNotFoundError(f"未找到映射表文件: {mapping_path}")
    wb = openpyxl.load_workbook(mapping_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    grid: List[List[str]] = []
    for r in range(1, 11):
        row_vals: List[str] = []
        for c in range(1, 11):
            v = ws.cell(row=r, column=c).value
            if v is None:
                raise ValueError(f"映射表存在空单元格: r{r} c{c}")
            row_vals.append(str(v).strip().replace("_", "-"))
        grid.append(row_vals)
    return grid


def build_gray_preview(image_bgr: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=1.2, tileGridSize=(8, 8))
    local = clahe.apply(gray)
    background = cv2.GaussianBlur(local, (0, 0), 5.0)
    detail = cv2.subtract(local, background)
    detail = cv2.normalize(detail, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
    return cv2.addWeighted(local, 0.82, detail, 0.18, 0.0)


def build_structure_mask(image_bgr: np.ndarray) -> np.ndarray:
    gray_local = build_gray_preview(image_bgr)
    background = cv2.GaussianBlur(gray_local, (0, 0), 11.0)
    highpass = cv2.subtract(gray_local, background)
    highpass = cv2.normalize(highpass, None, 0, 255, cv2.NORM_MINMAX)
    tophat = cv2.morphologyEx(gray_local, cv2.MORPH_TOPHAT, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (19, 19)))
    grad_x = cv2.Sobel(gray_local, cv2.CV_32F, 1, 0, ksize=3)
    grad_y = cv2.Sobel(gray_local, cv2.CV_32F, 0, 1, ksize=3)
    grad_mag = cv2.magnitude(grad_x, grad_y)
    grad_mag = cv2.GaussianBlur(grad_mag, (3, 3), 0)
    grad_mag = cv2.normalize(grad_mag, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
    _, m1 = cv2.threshold(highpass.astype(np.uint8), 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    _, m2 = cv2.threshold(tophat, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    _, m3 = cv2.threshold(grad_mag, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    mask = cv2.bitwise_or(cv2.bitwise_or(m1, m2), m3)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3)))
    mask = cv2.morphologyEx(mask, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5)))
    return mask


def collect_central_points(mask: np.ndarray) -> Tuple[np.ndarray, np.ndarray]:
    ys, xs = np.where(mask > 0)
    if len(xs) < 100:
        raise RuntimeError("亮结构太少，无法建立参考矩形。")
    center_x, center_y = float(np.mean(xs)), float(np.mean(ys))
    bbox_w, bbox_h = float(xs.max() - xs.min() + 1), float(ys.max() - ys.min() + 1)
    radius = 0.22 * max(bbox_w, bbox_h)
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(mask, connectivity=8)
    selected_mask = np.zeros_like(mask)
    points = []
    for idx in range(1, num_labels):
        if int(stats[idx, cv2.CC_STAT_AREA]) < 20:
            continue
        cx, cy = centroids[idx]
        if float(np.hypot(cx - center_x, cy - center_y)) > radius:
            continue
        ys_i, xs_i = np.where(labels == idx)
        if len(xs_i) == 0:
            continue
        selected_mask[ys_i, xs_i] = 255
        points.append(np.column_stack([xs_i.astype(np.float32), ys_i.astype(np.float32)]))
    if not points:
        raise RuntimeError("中心参考矩形提取失败。")
    return np.vstack(points), selected_mask


def normalize_angle_to_45(angle_deg: float) -> float:
    while angle_deg <= -45.0:
        angle_deg += 90.0
    while angle_deg > 45.0:
        angle_deg -= 90.0
    return angle_deg


def fit_reference_rectangle(points: np.ndarray) -> Tuple[np.ndarray, float]:
    hull = cv2.convexHull(points.reshape(-1, 1, 2))
    rect = cv2.minAreaRect(hull)
    box = cv2.boxPoints(rect).astype(np.float32)
    best_angle, best_abs = None, 1e9
    for idx in range(4):
        p1, p2 = box[idx], box[(idx + 1) % 4]
        angle = float(np.degrees(np.arctan2(float(p2[1] - p1[1]), float(p2[0] - p1[0]))))
        angle = normalize_angle_to_45(angle)
        if abs(angle) < best_abs:
            best_abs, best_angle = abs(angle), angle
    if best_angle is None:
        raise RuntimeError("参考矩形角度计算失败。")
    return np.int32(np.round(box)), float(best_angle)


def measure_side_angle(image_bgr: np.ndarray) -> Tuple[float, np.ndarray, np.ndarray, np.ndarray]:
    mask = build_structure_mask(image_bgr)
    points, selected_mask = collect_central_points(mask)
    box, angle_deg = fit_reference_rectangle(points)
    return angle_deg, mask, selected_mask, box


def estimate_dark_border_value(image: np.ndarray) -> Tuple[int, int, int]:
    h, w = image.shape[:2]
    bw, bh = max(8, w // 30), max(8, h // 30)
    strips = [image[:bh, :, :].reshape(-1, 3), image[-bh:, :, :].reshape(-1, 3), image[:, :bw, :].reshape(-1, 3), image[:, -bw:, :].reshape(-1, 3)]
    value = np.percentile(np.vstack(strips), 15, axis=0)
    return tuple(int(np.clip(v, 0, 255)) for v in value)


def rotate_keep_size(image: np.ndarray, delta_deg: float, border_value: Tuple[int, int, int]) -> np.ndarray:
    h, w = image.shape[:2]
    matrix = cv2.getRotationMatrix2D((w / 2.0, h / 2.0), delta_deg, 1.0)
    return cv2.warpAffine(image, matrix, (w, h), flags=cv2.INTER_LINEAR, borderMode=cv2.BORDER_CONSTANT, borderValue=border_value)


def refine_rotation_with_verification(image_bgr: np.ndarray, initial_angle_deg: float, border_value: Tuple[int, int, int]) -> Tuple[np.ndarray, float, float]:
    candidates = []
    for signed_angle in (-initial_angle_deg, initial_angle_deg):
        rotated = rotate_keep_size(image_bgr, signed_angle, border_value)
        try:
            residual, _, _, _ = measure_side_angle(rotated)
        except Exception:
            residual = 1e9
        candidates.append((abs(residual), rotated, signed_angle, residual))
    candidates.sort(key=lambda item: item[0])
    best_abs, best_image, applied_rotation_deg, residual = candidates[0]
    for _ in range(5):
        if best_abs <= 0.35:
            break
        updated = rotate_keep_size(best_image, -residual, border_value)
        try:
            new_residual, _, _, _ = measure_side_angle(updated)
        except Exception:
            break
        if abs(new_residual) >= best_abs:
            break
        best_image, applied_rotation_deg, residual, best_abs = updated, applied_rotation_deg - residual, new_residual, abs(new_residual)
    return best_image, applied_rotation_deg, residual


def build_color_invariant_mask(image_bgr: np.ndarray) -> np.ndarray:
    lab = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2LAB)
    a = lab[:, :, 1].astype(np.float32) - 128.0
    b = lab[:, :, 2].astype(np.float32) - 128.0
    chroma = np.sqrt(a * a + b * b)
    chroma = cv2.GaussianBlur(chroma, (3, 3), 0)
    chroma_u8 = (255 * (chroma / (np.max(chroma) + 1e-6))).astype(np.uint8)
    _, mask_chroma = cv2.threshold(chroma_u8, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    gray = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    gray_enh = clahe.apply(gray)
    tophat = cv2.morphologyEx(gray_enh, cv2.MORPH_TOPHAT, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (21, 21)))
    mask_gray = cv2.adaptiveThreshold(tophat, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, -2)
    mask = mask_chroma if int(np.count_nonzero(mask_chroma)) >= 3000 else cv2.bitwise_or(mask_chroma, mask_gray)
    return cv2.morphologyEx(mask, cv2.MORPH_CLOSE, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3)), iterations=1)


def detect_rectangles(image_bgr: np.ndarray, mask: np.ndarray, min_area: float = 180.0, max_area: float = 1800.0, max_aspect_ratio: float = 1.9) -> List[RectInfo]:
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    rects: List[RectInfo] = []
    for cnt in contours:
        area = float(cv2.contourArea(cnt))
        if area < min_area or area > max_area:
            continue
        rect = cv2.minAreaRect(cnt)
        rw, rh = rect[1]
        if rw < 1.0 or rh < 1.0:
            continue
        aspect = max(rw, rh) / min(rw, rh)
        if aspect > max_aspect_ratio:
            continue
        x, y, w, h = cv2.boundingRect(cnt)
        box = cv2.boxPoints(rect)
        rects.append({"bbox": (int(x), int(y), int(w), int(h)), "center": (float(x + w / 2.0), float(y + h / 2.0)), "area": area, "aspect": float(aspect), "rotated_box": np.int32(np.round(box)).tolist()})
    if len(rects) > 120:
        med_area = float(np.median([r["area"] for r in rects]))
        for r in rects:
            r["quality"] = abs(r["area"] - med_area) / max(med_area, 1e-6) + (r["aspect"] - 1.0) * 0.7
        rects = sorted(rects, key=lambda x: float(x["quality"]))[:120]
    return rects


def assign_labels(rects: List[RectInfo], mapping_grid_10x10: List[List[str]]) -> List[Dict[str, Any]]:
    rects_use = sorted(rects, key=lambda r: float(r.get("area", 0.0)), reverse=True)[:100]
    if len(rects_use) < 100:
        raise RuntimeError(f"可用于编号的矩形不足100个，当前: {len(rects_use)}")
    centers = np.array([r["center"] for r in rects_use], dtype=np.float32)
    xs = np.sort(centers[:, 0]); ys = np.sort(centers[:, 1])
    x_centers = [float(np.mean(xs[i * 10 : (i + 1) * 10])) for i in range(10)]
    y_centers = [float(np.mean(ys[i * 10 : (i + 1) * 10])) for i in range(10)]
    grid_points = [(rr, cc, x_centers[cc], y_centers[rr]) for rr in range(10) for cc in range(10)]
    pairs: List[Tuple[float, int, int]] = []
    for i, r in enumerate(rects_use):
        cx, cy = r["center"]
        for j, (_rr, _cc, gx, gy) in enumerate(grid_points):
            pairs.append((((float(cx) - gx) ** 2 + (float(cy) - gy) ** 2), i, j))
    pairs.sort(key=lambda x: x[0])
    used_rect, used_grid, assign = set(), set(), {}
    for _d, i, j in pairs:
        if i in used_rect or j in used_grid:
            continue
        used_rect.add(i); used_grid.add(j); assign[i] = j
        if len(assign) == 100:
            break
    if len(assign) < 100:
        raise RuntimeError(f"映射编号唯一匹配失败，已匹配: {len(assign)}")
    mapped = []
    for i, r in enumerate(rects_use):
        rr, cc, _gx, _gy = grid_points[assign[i]]
        label = mapping_grid_10x10[rr][cc]
        mapped.append({"label": label, "group_id": int(label.split("-")[0]), "rect_id": int(label.split("-")[1]), "center": (float(r["center"][0]), float(r["center"][1])), "bbox": tuple(int(v) for v in r["bbox"]), "rotated_box": r["rotated_box"]})
    mapped.sort(key=lambda x: (x["group_id"], x["rect_id"]))
    return mapped


def draw_results_mapped(image_bgr: np.ndarray, mapped_rects: List[Dict[str, Any]]) -> np.ndarray:
    out = image_bgr.copy()
    for m in mapped_rects:
        box = np.array(m["rotated_box"], dtype=np.int32).reshape((-1, 1, 2))
        cv2.polylines(out, [box], isClosed=True, color=(0, 0, 255), thickness=2)
        box_pts = np.array(m["rotated_box"], dtype=np.int32)
        x, y = int(np.min(box_pts[:, 0])), int(np.min(box_pts[:, 1]))
        cv2.putText(out, m["label"], (x, max(18, y - 4)), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 0, 255), 1, cv2.LINE_AA)
    return out


def build_overview_panel(original: np.ndarray, aligned: np.ndarray, mask: np.ndarray, final: np.ndarray) -> np.ndarray:
    h, w = original.shape[:2]
    half_w, half_h = w // 2, h // 2
    p1 = cv2.resize(original, (half_w, half_h), interpolation=cv2.INTER_AREA)
    p2 = cv2.resize(aligned, (half_w, half_h), interpolation=cv2.INTER_AREA)
    p3 = cv2.cvtColor(mask, cv2.COLOR_GRAY2BGR); p3 = cv2.resize(p3, (half_w, half_h), interpolation=cv2.INTER_NEAREST)
    p4 = cv2.resize(final, (half_w, half_h), interpolation=cv2.INTER_AREA)
    return np.vstack([np.hstack([p1, p2]), np.hstack([p3, p4])])


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="GBA1 standalone")
    parser.add_argument("--input", default=r"F:\研究生\项目\MRC_progress\mrc_test.png")
    parser.add_argument("--output", default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "result_gba1"))
    parser.add_argument("--mapping", default=r"F:\研究生\项目\MRC_progress\映射表.xlsx")
    args = parser.parse_args()

    image = imread_unicode(args.input)
    if image is None:
        raise FileNotFoundError(f"无法读取图像: {args.input}")
    mapping_grid = load_mapping_grid_10x10(args.mapping)
    angle_deg, _angle_mask, _selected_mask, _box = measure_side_angle(image)
    aligned, _, _ = refine_rotation_with_verification(image, angle_deg, estimate_dark_border_value(image))
    rect_mask = build_color_invariant_mask(aligned)
    rect_candidates = detect_rectangles(aligned, rect_mask)
    mapped_rects = assign_labels(rect_candidates, mapping_grid)
    final_img = draw_results_mapped(aligned, mapped_rects)
    overview = build_overview_panel(image, aligned, rect_mask, final_img)

    os.makedirs(args.output, exist_ok=True)
    stem = os.path.splitext(os.path.basename(args.input))[0]
    ext = os.path.splitext(args.input)[1] or ".png"
    out_a = os.path.join(args.output, f"{stem}_a{ext}")
    out_labels = os.path.join(args.output, f"{stem}_labels{ext}")
    out_ov = os.path.join(args.output, f"{stem}_ov{ext}")
    for p, img in [(out_a, aligned), (out_labels, final_img), (out_ov, overview)]:
        if not imwrite_unicode(p, img):
            raise IOError(f"保存失败: {p}")

    print(f"[OK] {stem}")

